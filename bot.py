import logging
import sqlite3
import os
import asyncio
from datetime import datetime, timedelta
from threading import Thread

from flask import Flask
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
    JobQueue,
)

# ==================== FLASK HTTP ENDPOINT (для UptimeRobot) ====================
flask_app = Flask(__name__)

@flask_app.route("/")
def health_check():
    return "✅ Бот работает!", 200

@flask_app.route("/keepalive")
def keepalive():
    return "OK", 200

def run_flask():
    flask_app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))

# ==================== ЛОГИРОВАНИЕ (СРАЗУ!) ====================
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ==================== НАСТРОЙКИ ====================
# Получаем токен из переменной окружения (без кавычек!)
raw_token = os.environ.get("TOKEN", "")
TOKEN = raw_token.strip().strip('"').strip("'")

if not TOKEN:
    logger.error("❌ TOKEN не задан! Установи переменную окружения TOKEN в Render Dashboard")
    raise ValueError("TOKEN не задан")

logger.info(f"🔑 TOKEN загружен, длина: {len(TOKEN)}")

# Пути к файлам
DB_PATH = os.path.join(os.path.dirname(__file__), "shifts.db")

# ==================== СОСТОЯНИЯ ДЛЯ ConversationHandler ====================
SELECT_BREAK = 1
SELECT_TEMPLATE = 2
SETTINGS_MENU = 3

# ==================== БАЗА ДАННЫХ ====================
def init_db():
    """Создаём таблицы если их нет"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Таблица пользователей
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            username TEXT,
            first_name TEXT,
            created_at TEXT,
            week_start TEXT,
            week_hours REAL DEFAULT 0,
            two_week_hours REAL DEFAULT 0
        )
    """)

    # Таблица смен
    c.execute("""
        CREATE TABLE IF NOT EXISTS shifts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            start_time TEXT,
            end_time TEXT,
            total_work REAL,
            total_break REAL,
            breaks_count INTEGER,
            status TEXT,
            violation TEXT,
            template_name TEXT,
            FOREIGN KEY (user_id) REFERENCES users(user_id)
        )
    """)

    # Таблица перерывов
    c.execute("""
        CREATE TABLE IF NOT EXISTS breaks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shift_id INTEGER,
            start_time TEXT,
            end_time TEXT,
            duration INTEGER,
            FOREIGN KEY (shift_id) REFERENCES shifts(id)
        )
    """)

    # Таблица нарушений
    c.execute("""
        CREATE TABLE IF NOT EXISTS violations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            shift_id INTEGER,
            type TEXT,
            description TEXT,
            time TEXT,
            FOREIGN KEY (user_id) REFERENCES users(user_id)
        )
    """)

    conn.commit()
    conn.close()
    logger.info("База данных инициализирована")


def get_conn():
    """Получить соединение с БД"""
    return sqlite3.connect(DB_PATH)


# ==================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ====================
def now_str():
    """Текущее время в строке (Минск, UTC+3)"""
    import pytz
    minsk_tz = pytz.timezone('Europe/Minsk')
    return datetime.now(minsk_tz).strftime("%Y-%m-%d %H:%M:%S")


def str_to_dt(s):
    """Строку в datetime"""
    return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")


def format_duration(minutes):
    """Форматировать минуты в часы и минуты"""
    h = int(minutes // 60)
    m = int(minutes % 60)
    if h > 0:
        return f"{h}ч {m}мин"
    return f"{m}мин"


def format_time(dt):
    """Красивый формат времени (Минск, UTC+3)"""
    import pytz
    if dt.tzinfo is None:
        minsk_tz = pytz.timezone('Europe/Minsk')
        dt = minsk_tz.localize(dt)
    return dt.strftime("%d.%m.%Y %H:%M")


# ==================== ПРОВЕРКА ЛИМИТОВ ====================
def check_week_limits(user_id):
    """Проверить недельные лимиты"""
    conn = get_conn()
    c = conn.cursor()

    c.execute("SELECT week_start, week_hours, two_week_hours FROM users WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()

    if not row:
        return None

    week_start, week_hours, two_week_hours = row

    # Если неделя не начата или прошло больше 7 дней — сброс
    if week_start:
        ws = str_to_dt(week_start)
        if datetime.now() - ws > timedelta(days=7):
            # Начинаем новую неделю
            reset_week(user_id)
            return {"week_hours": 0, "two_week_hours": two_week_hours, "new_week": True}

    return {
        "week_hours": week_hours or 0,
        "two_week_hours": two_week_hours or 0,
        "new_week": False
    }


def reset_week(user_id):
    """Сбросить недельный счётчик"""
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT week_hours FROM users WHERE user_id=?", (user_id,))
    old_week = c.fetchone()
    old_week_hours = old_week[0] if old_week else 0

    c.execute("""
        UPDATE users 
        SET week_start=?, week_hours=0, two_week_hours=two_week_hours-?+?
        WHERE user_id=?
    """, (now_str(), old_week_hours, 0, user_id))
    conn.commit()
    conn.close()


def add_work_hours(user_id, hours):
    """Добавить часы в недельный счётчик"""
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        UPDATE users 
        SET week_hours=week_hours+?, two_week_hours=two_week_hours+?
        WHERE user_id=?
    """, (hours, hours, user_id))
    conn.commit()
    conn.close()


# ==================== ШАБЛОНЫ СМЕН ====================
TEMPLATES = {
    "day": {"name": "🌅 Дневная", "hours": 8, "desc": "8 часов (4+4 с перерывом)"},
    "night": {"name": "🌙 Ночная", "hours": 10, "desc": "10 часов (5+5 с перерывом)"},
    "short": {"name": "⚡ Короткая", "hours": 6, "desc": "6 часов (3+3 с перерывом)"},
    "long": {"name": "🔥 Длинная", "hours": 11, "desc": "11 часов (5.5+5.5) — макс!"},
    "custom": {"name": "⚙️ Своя", "hours": 0, "desc": "Настроить вручную"},
}


# ==================== ГЛАВНОЕ МЕНЮ ====================
def main_menu_keyboard():
    """Клавиатура главного меню"""
    keyboard = [
        [InlineKeyboardButton("▶️ Начать смену", callback_data="start_shift")],
        [InlineKeyboardButton("⏸️ Перерыв", callback_data="take_break"),
         InlineKeyboardButton("▶️ На линии", callback_data="resume_work")],
        [InlineKeyboardButton("⏹️ Завершить смену", callback_data="end_shift")],
        [InlineKeyboardButton("📊 Статистика", callback_data="stats"),
         InlineKeyboardButton("📁 Экспорт", callback_data="export")],
        [InlineKeyboardButton("⚙️ Настройки", callback_data="settings")],
        [InlineKeyboardButton("❓ Помощь", callback_data="help")],
    ]
    return InlineKeyboardMarkup(keyboard)


def break_keyboard():
    """Клавиатура выбора перерыва"""
    keyboard = [
        [InlineKeyboardButton("⏱️ 15 минут", callback_data="break_15")],
        [InlineKeyboardButton("⏱️ 30 минут", callback_data="break_30")],
        [InlineKeyboardButton("❌ Отмена", callback_data="cancel_break")],
    ]
    return InlineKeyboardMarkup(keyboard)


def template_keyboard():
    """Клавиатура шаблонов"""
    keyboard = []
    for key, tmpl in TEMPLATES.items():
        keyboard.append([InlineKeyboardButton(f"{tmpl['name']} — {tmpl['desc']}", callback_data=f"template_{key}")])
    keyboard.append([InlineKeyboardButton("❌ Отмена", callback_data="cancel_template")])
    return InlineKeyboardMarkup(keyboard)


# ==================== КОМАНДЫ ====================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда /start — приветствие и регистрация"""
    user = update.effective_user
    user_id = user.id

    # Регистрируем пользователя
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT user_id FROM users WHERE user_id=?", (user_id,))
    if not c.fetchone():
        c.execute("""
            INSERT INTO users (user_id, username, first_name, created_at, week_start)
            VALUES (?, ?, ?, ?, ?)
        """, (user_id, user.username, user.first_name, now_str(), now_str()))
        conn.commit()
        logger.info(f"Новый пользователь: {user_id} ({user.first_name})")
    conn.close()

    text = (
        f"👋 Привет, {user.first_name}!\n\n"
        f"🚕 Я бот для контроля режима труда и отдыха (РТО) таксиста.\n\n"
        f"📋 *Что я умею:*\n"
        f"• Отслеживать время смены\n"
        f"• Напоминать о перерывах\n"
        f"• Следить за лимитами (9ч/12ч/40ч/56ч/90ч)\n"
        f"• Вести статистику и историю\n"
        f"• Экспортировать данные в Excel\n\n"
        f"⚠️ *Важно:* Убедись, что уведомления в Telegram включены!\n\n"
        f"Выбери действие ниже 👇"
    )

    await update.message.reply_text(
        text,
        parse_mode="Markdown",
        reply_markup=main_menu_keyboard()
    )


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда /help — помощь"""
    text = (
        "📖 *Памятка РТО (Беларусь)*\n\n"
        "🔴 *Главные цифры:*\n"
        "• 4 часа → перерыв 30 мин (можно 2×15)\n"
        "• 9 часов → лимит чистого вождения\n"
        "• 12 часов → максимум смены\n\n"
        "💤 *Отдых:*\n"
        "• 11 часов — минимум между сменами\n"
        "• 42 часа — выходные подряд\n"
        "• 4 выходных в месяц\n\n"
        "📊 *Недельные лимиты:*\n"
        "• 40ч — норма\n"
        "• 56ч — потолок за неделю\n"
        "• 90ч — максимум за 2 недели\n\n"
        "⚠️ *Штрафы:* ст. 18.26 ч.2 КоАП РБ\n"
        "Повторное нарушение 42-часового отдыха = исключение из реестра!"
    )

    if update.callback_query:
        await update.callback_query.edit_message_text(text, parse_mode="Markdown", reply_markup=main_menu_keyboard())
    else:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=main_menu_keyboard())


# ==================== ОБРАБОТКА КНОПОК ====================
async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик всех кнопок"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    data = query.data

    # ===== НАЧАТЬ СМЕНУ =====
    if data == "start_shift":
        await handle_start_shift(query, context, user_id)

    # ===== ВЫБОР ШАБЛОНА =====
    elif data.startswith("template_"):
        template_key = data.replace("template_", "")
        await handle_template_selected(query, context, user_id, template_key)

    elif data == "cancel_template":
        await query.edit_message_text("❌ Смена не начата.", reply_markup=main_menu_keyboard())

    # ===== ПЕРЕРЫВ =====
    elif data == "take_break":
        await handle_take_break(query, context, user_id)

    elif data.startswith("break_"):
        minutes = int(data.replace("break_", ""))
        await handle_break_start(query, context, user_id, minutes)

    elif data == "cancel_break":
        await query.edit_message_text("❌ Перерыв отменён.", reply_markup=main_menu_keyboard())

    # ===== НА ЛИНИИ (закончить перерыв) =====
    elif data == "resume_work":
        await handle_resume_work(query, context, user_id)

    # ===== ЗАВЕРШИТЬ СМЕНУ =====
    elif data == "end_shift":
        await handle_end_shift(query, context, user_id)

    # ===== СТАТИСТИКА =====
    elif data == "stats":
        await handle_stats(query, context, user_id)

    # ===== ЭКСПОРТ =====
    elif data == "export":
        await handle_export(query, context, user_id)

    # ===== НАСТРОЙКИ =====
    elif data == "settings":
        await handle_settings(query, context, user_id)

    # ===== ПОМОЩЬ =====
    elif data == "help":
        await help_command(update, context)

    # ===== НАЗАД В МЕНЮ =====
    elif data == "back_menu":
        await query.edit_message_text(
            "🚕 Главное меню. Выбери действие:",
            reply_markup=main_menu_keyboard()
        )


# ==================== ЛОГИКА СМЕНЫ ====================
async def handle_start_shift(query, context, user_id):
    """Начать смену — проверки и выбор шаблона"""
    conn = get_conn()
    c = conn.cursor()

    # Проверяем, нет ли активной смены
    c.execute("""
        SELECT id, start_time FROM shifts 
        WHERE user_id=? AND status='active' 
        ORDER BY id DESC LIMIT 1
    """, (user_id,))
    active = c.fetchone()

    if active:
        conn.close()
        await query.edit_message_text(
            "⚠️ У тебя уже есть активная смена!\n"
            f"Начата: {format_time(str_to_dt(active[1]))}\n\n"
            "Завершите текущую смену перед началом новой.",
            reply_markup=main_menu_keyboard()
        )
        return

    # Проверяем отдых между сменами (11 часов)
    c.execute("""
        SELECT end_time FROM shifts 
        WHERE user_id=? AND status='completed' 
        ORDER BY id DESC LIMIT 1
    """, (user_id,))
    last_shift = c.fetchone()
    conn.close()

    if last_shift and last_shift[0]:
        last_end = str_to_dt(last_shift[0])
        rest_hours = (datetime.now() - last_end).total_seconds() / 3600

        if rest_hours < 11:
            remaining = 11 - rest_hours
            await query.edit_message_text(
                f"🚫 *Нельзя начать смену!*\n\n"
                f"Последняя смена закончена: {format_time(last_end)}\n"
                f"Прошло: {rest_hours:.1f}ч из 11ч минимума\n"
                f"Осталось отдыхать: *{remaining:.1f}ч*\n\n"
                f"⏰ Можно начать после: {format_time(last_end + timedelta(hours=11))}",
                parse_mode="Markdown",
                reply_markup=main_menu_keyboard()
            )
            return

    # Проверяем недельные лимиты
    limits = check_week_limits(user_id)
    if limits and limits["week_hours"] >= 56:
        await query.edit_message_text(
            "🚫 *Недельный лимит исчерпан!*\n"
            f"Уже наработано: *{limits['week_hours']:.1f}ч* / 56ч\n"
            "Начните новую неделю отдыхом 42 часа.",
            parse_mode="Markdown",
            reply_markup=main_menu_keyboard()
        )
        return

    if limits and limits["two_week_hours"] >= 90:
        await query.edit_message_text(
            "🚫 *Двухнедельный лимит исчерпан!*\n"
            f"Уже наработано: *{limits['two_week_hours']:.1f}ч* / 90ч\n"
            "Обязательный отдых перед продолжением.",
            parse_mode="Markdown",
            reply_markup=main_menu_keyboard()
        )
        return

    # Показываем шаблоны
    text = "📋 Выбери шаблон смены:\n\n"
    for key, tmpl in TEMPLATES.items():
        emoji = "🌅" if key == "day" else "🌙" if key == "night" else "⚡" if key == "short" else "🔥" if key == "long" else "⚙️"
        text += f"{emoji} *{tmpl['name']}* — {tmpl['desc']}\n"

    await query.edit_message_text(text, parse_mode="Markdown", reply_markup=template_keyboard())


async def handle_template_selected(query, context, user_id, template_key):
    """Шаблон выбран — начинаем смену"""
    template = TEMPLATES.get(template_key, TEMPLATES["day"])

    conn = get_conn()
    c = conn.cursor()

    start_time = now_str()
    c.execute("""
        INSERT INTO shifts (user_id, start_time, total_work, total_break, breaks_count, status, template_name)
        VALUES (?, ?, 0, 0, 0, 'active', ?)
    """, (user_id, start_time, template["name"]))

    shift_id = c.lastrowid
    conn.commit()
    conn.close()

    # Сохраняем ID смены в контексте
    context.user_data["shift_id"] = shift_id
    context.user_data["shift_start"] = start_time
    context.user_data["template"] = template_key

    # Устанавливаем напоминания
    job_queue = context.application.job_queue

    # Напоминание о перерыве за 10 мин до 4 часов (3ч 50мин = 13800 сек)
    job_queue.run_once(
        reminder_break_soon,
        when=13800,
        chat_id=query.message.chat_id,
        user_id=user_id,
        name=f"remind_break_{user_id}",
        data={"shift_id": shift_id, "type": "break_soon"}
    )

    # Жёсткое напоминание на 4 часах (14400 сек)
    job_queue.run_once(
        reminder_break_now,
        when=14400,
        chat_id=query.message.chat_id,
        user_id=user_id,
        name=f"break_now_{user_id}",
        data={"shift_id": shift_id, "type": "break_now"}
    )

    # Предупреждение на 9 часах (32400 сек)
    job_queue.run_once(
        reminder_9_hours,
        when=32400,
        chat_id=query.message.chat_id,
        user_id=user_id,
        name=f"warn_9h_{user_id}",
        data={"shift_id": shift_id, "type": "warn_9h"}
    )

    # Предупреждение на 11 часах (39600 сек)
    job_queue.run_once(
        reminder_11_hours,
        when=39600,
        chat_id=query.message.chat_id,
        user_id=user_id,
        name=f"warn_11h_{user_id}",
        data={"shift_id": shift_id, "type": "warn_11h"}
    )

    # Автозавершение на 12 часах (43200 сек)
    job_queue.run_once(
        auto_end_shift,
        when=43200,
        chat_id=query.message.chat_id,
        user_id=user_id,
        name=f"auto_end_{user_id}",
        data={"shift_id": shift_id, "type": "auto_end"}
    )

    text = (
        f"✅ *Смена начата!*\n\n"
        f"🕐 Начало: *{format_time(datetime.now())}*\n"
        f"📋 Шаблон: *{template['name']}*\n\n"
        f"⏰ *Напоминания установлены:*\n"
        f"• 3ч 50мин — предупреждение о перерыве\n"
        f"• 4ч 00мин — время перерыва!\n"
        f"• 9ч 00мин — предупреждение о лимите\n"
        f"• 11ч 00мин — строгое предупреждение\n"
        f"• 12ч 00мин — *автозавершение смены*\n\n"
        f"💡 Не забудь включить уведомления в Telegram!"
    )

    await query.edit_message_text(text, parse_mode="Markdown", reply_markup=main_menu_keyboard())


# ==================== НАПОМИНАНИЯ (Job Queue) ====================
async def reminder_break_soon(context: ContextTypes.DEFAULT_TYPE):
    """Напоминание за 10 мин до перерыва"""
    job = context.job
    chat_id = job.chat_id
    user_id = job.user_id

    await context.bot.send_message(
        chat_id=chat_id,
        text=(
            "⏰ *Через 10 минут перерыв!*\n\n"
            "Ты работаешь уже 3ч 50мин.\n"
            "Подготовься: найди место, закончи заказ.\n\n"
            "Нажми *⏸️ Перерыв* когда будешь готов."
        ),
        parse_mode="Markdown",
        reply_markup=main_menu_keyboard()
    )


async def reminder_break_now(context: ContextTypes.DEFAULT_TYPE):
    """Время перерыва!"""
    job = context.job
    chat_id = job.chat_id
    user_id = job.user_id

    await context.bot.send_message(
        chat_id=chat_id,
        text=(
            "🔴 *ВРЕМЯ ПЕРЕРЫВА!*\n\n"
            "Ты работаешь уже *4 часа*.\n"
            "Обязательный перерыв 30 минут (можно 2×15).\n\n"
            "⚠️ Без перерыва — нарушение РТО!"
        ),
        parse_mode="Markdown",
        reply_markup=main_menu_keyboard()
    )


async def reminder_9_hours(context: ContextTypes.DEFAULT_TYPE):
    """Предупреждение на 9 часах"""
    job = context.job
    chat_id = job.chat_id

    await context.bot.send_message(
        chat_id=chat_id,
        text=(
            "⚠️ *ЛИМИТ 9 ЧАСОВ!*\n\n"
            "Ты приближаешься к максимуму чистого вождения.\n"
            "Рекомендуется завершить смену в ближайшее время.\n\n"
            "📊 Дальше — только предупреждения о нарушениях."
        ),
        parse_mode="Markdown",
        reply_markup=main_menu_keyboard()
    )


async def reminder_11_hours(context: ContextTypes.DEFAULT_TYPE):
    """Строгое предупреждение на 11 часах"""
    job = context.job
    chat_id = job.chat_id

    await context.bot.send_message(
        chat_id=chat_id,
        text=(
            "🚨 *НАРУШЕНИЕ РТО!*\n\n"
            "Ты работаешь уже *11 часов*!\n\n"
            "⚠️ *Штраф по ст. 18.26 ч.2 КоАП РБ*\n"
            "• Превышение времени вождения\n"
            "• Пропуск обязательных перерывов\n\n"
            "🛑 *Смена автоматически завершится через 1 час!*"
        ),
        parse_mode="Markdown",
        reply_markup=main_menu_keyboard()
    )


async def auto_end_shift(context: ContextTypes.DEFAULT_TYPE):
    """Автозавершение смены на 12 часах"""
    job = context.job
    chat_id = job.chat_id
    user_id = job.user_id
    shift_id = job.data["shift_id"]

    conn = get_conn()
    c = conn.cursor()

    # Проверяем, не завершена ли уже
    c.execute("SELECT status, start_time FROM shifts WHERE id=?", (shift_id,))
    row = c.fetchone()
    if not row or row[0] != "active":
        conn.close()
        return

    start_time = str_to_dt(row[1])
    end_time = datetime.now()
    total_work = (end_time - start_time).total_seconds() / 3600

    # Завершаем смену
    c.execute("""
        UPDATE shifts 
        SET end_time=?, total_work=?, status='completed', violation='auto_12h'
        WHERE id=?
    """, (now_str(), total_work, shift_id))

    # Записываем нарушение
    c.execute("""
        INSERT INTO violations (user_id, shift_id, type, description, time)
        VALUES (?, ?, 'auto_end', 'Смена автоматически завершена на 12 часах', ?)
    """, (user_id, shift_id, now_str()))

    conn.commit()
    conn.close()

    # Добавляем часы в недельный счёт
    add_work_hours(user_id, total_work)

    await context.bot.send_message(
        chat_id=chat_id,
        text=(
            "🛑 *СМЕНА АВТОМАТИЧЕСКИ ЗАВЕРШЕНА!*\n\n"
            f"⏱️ Общее время: *{total_work:.1f}ч*\n"
            f"🕐 Завершена: *{format_time(end_time)}*\n\n"
            "⚠️ *Нарушение зафиксировано:*\n"
            "Превышение максимальной длительности смены (12ч)\n"
            "Записано в историю нарушений.\n\n"
            "💤 *Обязательный отдых 11 часов перед следующей сменой!*"
        ),
        parse_mode="Markdown",
        reply_markup=main_menu_keyboard()
    )


# ==================== ПЕРЕРЫВ ====================
async def handle_take_break(query, context, user_id):
    """Начать перерыв"""
    conn = get_conn()
    c = conn.cursor()

    # Находим активную смену
    c.execute("""
        SELECT id, start_time, total_work, breaks_count FROM shifts 
        WHERE user_id=? AND status='active' 
        ORDER BY id DESC LIMIT 1
    """, (user_id,))
    shift = c.fetchone()
    conn.close()

    if not shift:
        await query.edit_message_text(
            "⚠️ Нет активной смены! Начни смену сначала.",
            reply_markup=main_menu_keyboard()
        )
        return

    shift_id, start_time, total_work, breaks_count = shift

    # Проверяем, не на перерыве ли уже
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        SELECT id FROM breaks 
        WHERE shift_id=? AND end_time IS NULL 
        ORDER BY id DESC LIMIT 1
    """, (shift_id,))
    active_break = c.fetchone()
    conn.close()

    if active_break:
        await query.edit_message_text(
            "⚠️ У тебя уже активный перерыв!\n"
            "Нажми *▶️ На линии* чтобы закончить его.",
            reply_markup=main_menu_keyboard()
        )
        return

    context.user_data["break_shift_id"] = shift_id

    await query.edit_message_text(
        "⏸️ Выбери длительность перерыва:\n\n"
        "• 15 мин — половина (потом ещё 15)\n"
        "• 30 мин — полный перерыв",
        reply_markup=break_keyboard()
    )


async def handle_break_start(query, context, user_id, minutes):
    """Начать перерыв выбранной длительности"""
    shift_id = context.user_data.get("break_shift_id")
    if not shift_id:
        await query.edit_message_text("❌ Ошибка: смена не найдена.", reply_markup=main_menu_keyboard())
        return

    break_start = now_str()

    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        INSERT INTO breaks (shift_id, start_time, duration)
        VALUES (?, ?, ?)
    """, (shift_id, break_start, minutes))

    break_id = c.lastrowid
    conn.commit()
    conn.close()

    context.user_data["break_id"] = break_id
    context.user_data["break_minutes"] = minutes

    # Устанавливаем таймер окончания перерыва
    job_queue = context.application.job_queue
    job_queue.run_once(
        reminder_break_end,
        when=minutes * 60,
        chat_id=query.message.chat_id,
        user_id=user_id,
        name=f"break_end_{user_id}",
        data={"break_id": break_id, "minutes": minutes}
    )

    end_time = datetime.now() + timedelta(minutes=minutes)

    await query.edit_message_text(
        f"⏸️ *Перерыв начат!*\n\n"
        f"⏱️ Длительность: *{minutes} минут*\n"
        f"🕐 Закончится в: *{format_time(end_time)}*\n\n"
        f"Я напомню, когда пора возвращаться на линию!",
        parse_mode="Markdown",
        reply_markup=main_menu_keyboard()
    )


async def reminder_break_end(context: ContextTypes.DEFAULT_TYPE):
    """Напоминание об окончании перерыва"""
    job = context.job
    chat_id = job.chat_id
    minutes = job.data["minutes"]

    await context.bot.send_message(
        chat_id=chat_id,
        text=(
            f"✅ *Перерыв закончен!* ({minutes} мин)\n\n"
            f"Пора возвращаться на линию.\n"
            f"Нажми *▶️ На линии* чтобы продолжить смену."
        ),
        parse_mode="Markdown",
        reply_markup=main_menu_keyboard()
    )


# ==================== НА ЛИНИИ (закончить перерыв) ====================
async def handle_resume_work(query, context, user_id):
    """Закончить перерыв, вернуться на линию"""
    conn = get_conn()
    c = conn.cursor()

    # Находим активную смену
    c.execute("""
        SELECT id FROM shifts 
        WHERE user_id=? AND status='active' 
        ORDER BY id DESC LIMIT 1
    """, (user_id,))
    shift = c.fetchone()

    if not shift:
        conn.close()
        await query.edit_message_text(
            "⚠️ Нет активной смены!",
            reply_markup=main_menu_keyboard()
        )
        return

    shift_id = shift[0]

    # Находим активный перерыв
    c.execute("""
        SELECT id, start_time, duration FROM breaks 
        WHERE shift_id=? AND end_time IS NULL 
        ORDER BY id DESC LIMIT 1
    """, (shift_id,))
    break_row = c.fetchone()

    if not break_row:
        conn.close()
        await query.edit_message_text(
            "⚠️ Нет активного перерыва!\n"
            "Ты уже на линии или смена не начата.",
            reply_markup=main_menu_keyboard()
        )
        return

    break_id, break_start, planned_duration = break_row
    break_end = now_str()

    # Считаем фактическую длительность перерыва
    actual_duration = (datetime.now() - str_to_dt(break_start)).total_seconds() / 60

    # Закрываем перерыв
    c.execute("""
        UPDATE breaks SET end_time=?, duration=? WHERE id=?
    """, (break_end, actual_duration, break_id))

    # Обновляем смену
    c.execute("""
        UPDATE shifts 
        SET total_break = total_break + ?, breaks_count = breaks_count + 1
        WHERE id=?
    """, (actual_duration, shift_id))

    conn.commit()
    conn.close()

    await query.edit_message_text(
        f"▶️ *Снова на линии!*\n\n"
        f"⏸️ Перерыв: *{format_duration(actual_duration)}*\n"
        f"(планировалось: {planned_duration} мин)\n\n"
        f"Продолжай работу! Следующее напоминание о перерыве через 4 часа.",
        parse_mode="Markdown",
        reply_markup=main_menu_keyboard()
    )


# ==================== ЗАВЕРШИТЬ СМЕНУ ====================
async def handle_end_shift(query, context, user_id):
    """Завершить смену вручную"""
    conn = get_conn()
    c = conn.cursor()

    c.execute("""
        SELECT id, start_time, total_break, breaks_count FROM shifts 
        WHERE user_id=? AND status='active' 
        ORDER BY id DESC LIMIT 1
    """, (user_id,))
    shift = c.fetchone()

    if not shift:
        conn.close()
        await query.edit_message_text(
            "⚠️ Нет активной смены для завершения!",
            reply_markup=main_menu_keyboard()
        )
        return

    shift_id, start_time, total_break, breaks_count = shift

    # Если есть активный перерыв — закрываем его
    c.execute("""
        SELECT id, start_time FROM breaks 
        WHERE shift_id=? AND end_time IS NULL 
        ORDER BY id DESC LIMIT 1
    """, (shift_id,))
    active_break = c.fetchone()

    if active_break:
        break_id, break_start = active_break
        actual_duration = (datetime.now() - str_to_dt(break_start)).total_seconds() / 60
        c.execute("UPDATE breaks SET end_time=?, duration=? WHERE id=?", (now_str(), actual_duration, break_id))
        total_break += actual_duration
        breaks_count += 1

    end_time = datetime.now()
    total_work = (end_time - str_to_dt(start_time)).total_seconds() / 3600
    pure_work = total_work - (total_break / 60)

    # Определяем нарушения
    violations = []
    if total_work > 12:
        violations.append("Превышение 12ч смены")
    if pure_work > 9:
        violations.append("Превышение 9ч чистого вождения")
    if total_break < 30 and total_work > 4:
        violations.append("Недостаточный перерыв (<30 мин)")

    violation_str = "; ".join(violations) if violations else None

    # Завершаем смену
    c.execute("""
        UPDATE shifts 
        SET end_time=?, total_work=?, total_break=?, breaks_count=?, status='completed', violation=?
        WHERE id=?
    """, (now_str(), total_work, total_break, breaks_count, violation_str, shift_id))

    # Записываем нарушения
    for v in violations:
        c.execute("""
            INSERT INTO violations (user_id, shift_id, type, description, time)
            VALUES (?, ?, 'manual_end', ?, ?)
        """, (user_id, shift_id, v, now_str()))

    conn.commit()
    conn.close()

    # Добавляем часы в недельный счёт
    add_work_hours(user_id, total_work)

    # Отменяем все запланированные напоминания
    job_queue = context.application.job_queue
    for job_name in [f"remind_break_{user_id}", f"break_now_{user_id}", f"warn_9h_{user_id}", f"warn_11h_{user_id}", f"auto_end_{user_id}"]:
        jobs = job_queue.get_jobs_by_name(job_name)
        for job in jobs:
            job.schedule_removal()

    # Формируем отчёт
    text = (
        f"⏹️ *Смена завершена!*\n\n"
        f"🕐 Начало: *{format_time(str_to_dt(start_time))}*\n"
        f"🕐 Конец: *{format_time(end_time)}*\n"
        f"⏱️ Общее время: *{total_work:.1f}ч*\n"
        f"⏸️ Перерывов: *{breaks_count}* (всего {format_duration(total_break)})\n"
        f"🚗 Чистое вождение: *{pure_work:.1f}ч*\n"
    )

    if violations:
        text += f"\n🚨 *Нарушения:*\n"
        for v in violations:
            text += f"• {v}\n"
        text += f"\n⚠️ Записано в историю нарушений."
    else:
        text += f"\n✅ *Нарушений нет!* Молодец!"

    text += f"\n\n💤 Следующую смену можно начать через *11 часов*."

    await query.edit_message_text(text, parse_mode="Markdown", reply_markup=main_menu_keyboard())


# ==================== СТАТИСТИКА ====================
async def handle_stats(query, context, user_id):
    """Показать статистику"""
    conn = get_conn()
    c = conn.cursor()

    # Сегодня
    today = datetime.now().strftime("%Y-%m-%d")
    c.execute("""
        SELECT COUNT(*), SUM(total_work), SUM(total_break) FROM shifts 
        WHERE user_id=? AND DATE(start_time)=?
    """, (user_id, today))
    today_stats = c.fetchone()

    # Неделя
    week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")
    c.execute("""
        SELECT COUNT(*), SUM(total_work), SUM(total_break) FROM shifts 
        WHERE user_id=? AND start_time >= ? AND status='completed'
    """, (user_id, week_ago))
    week_stats = c.fetchone()

    # Месяц
    month_ago = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S")
    c.execute("""
        SELECT COUNT(*), SUM(total_work) FROM shifts 
        WHERE user_id=? AND start_time >= ? AND status='completed'
    """, (user_id, month_ago))
    month_stats = c.fetchone()

    # Нарушения
    c.execute("SELECT COUNT(*) FROM violations WHERE user_id=?", (user_id,))
    violations_count = c.fetchone()[0] or 0

    # Недельные лимиты
    limits = check_week_limits(user_id)

    conn.close()

    today_count = today_stats[0] or 0
    today_work = today_stats[1] or 0
    today_break = today_stats[2] or 0

    week_count = week_stats[0] or 0
    week_work = week_stats[1] or 0

    month_count = month_stats[0] or 0
    month_work = month_stats[1] or 0

    text = (
        f"📊 *Твоя статистика*\n\n"
        f"📅 *Сегодня ({today}):*\n"
        f"• Смен: {today_count}\n"
        f"• Работы: {today_work:.1f}ч\n"
        f"• Перерывов: {format_duration(today_break)}\n\n"
        f"📆 *За 7 дней:*\n"
        f"• Смен: {week_count}\n"
        f"• Работы: {week_work:.1f}ч / 56ч\n"
        f"• Прогресс: {'🟢' if week_work < 40 else '🟡' if week_work < 56 else '🔴'} {min(week_work/56*100, 100):.0f}%\n\n"
        f"📈 *За 30 дней:*\n"
        f"• Смен: {month_count}\n"
        f"• Работы: {month_work:.1f}ч\n\n"
        f"🚨 *Нарушений всего:* {violations_count}\n"
    )

    if limits:
        text += f"\n📊 *Текущая неделя:*\n"
        text += f"• Наработано: {limits['week_hours']:.1f}ч / 56ч\n"
        text += f"• 2 недели: {limits['two_week_hours']:.1f}ч / 90ч\n"

    await query.edit_message_text(text, parse_mode="Markdown", reply_markup=main_menu_keyboard())


# ==================== ЭКСПОРТ В EXCEL ====================
async def handle_export(query, context, user_id):
    """Экспорт данных в Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    conn = get_conn()
    c = conn.cursor()

    # Получаем все смены
    c.execute("""
        SELECT id, start_time, end_time, total_work, total_break, breaks_count, status, violation, template_name
        FROM shifts WHERE user_id=? ORDER BY start_time DESC
    """, (user_id,))
    shifts = c.fetchall()

    # Получаем нарушения
    c.execute("""
        SELECT type, description, time FROM violations WHERE user_id=? ORDER BY time DESC
    """, (user_id,))
    violations = c.fetchall()

    conn.close()

    if not shifts:
        await query.edit_message_text(
            "📭 Пока нет данных для экспорта.\nНачни несколько смен!",
            reply_markup=main_menu_keyboard()
        )
        return

    # Создаём Excel
    wb = openpyxl.Workbook()

    # Лист 1: Смены
    ws1 = wb.active
    ws1.title = "Смены"

    headers = ["№", "Начало", "Конец", "Работа (ч)", "Перерыв (мин)", "Перерывов", "Статус", "Нарушения", "Шаблон"]
    ws1.append(headers)

    for header in ws1[1]:
        header.font = Font(bold=True, color="FFFFFF")
        header.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header.alignment = Alignment(horizontal="center")

    for shift in shifts:
        ws1.append([
            shift[0],
            shift[1],
            shift[2] or "—",
            round(shift[3], 2) if shift[3] else 0,
            round(shift[4], 1) if shift[4] else 0,
            shift[5] or 0,
            shift[6],
            shift[7] or "Нет",
            shift[8] or "—"
        ])

    # Автоширина
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column].width = adjusted_width

    # Лист 2: Нарушения
    ws2 = wb.create_sheet("Нарушения")
    ws2.append(["Тип", "Описание", "Время"])

    for header in ws2[1]:
        header.font = Font(bold=True, color="FFFFFF")
        header.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    for v in violations:
        ws2.append(list(v))

    for col in ws2.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws2.column_dimensions[column].width = min(max_length + 2, 60)

    # Сохраняем
    filename = f"rto_stats_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(os.path.dirname(__file__), filename)
    wb.save(filepath)

    # Отправляем файл
    with open(filepath, "rb") as f:
        await context.bot.send_document(
            chat_id=query.message.chat_id,
            document=f,
            caption=f"📁 *Экспорт статистики РТО*\n\n"
                    f"• Смен: {len(shifts)}\n"
                    f"• Нарушений: {len(violations)}\n"
                    f"• Дата: {format_time(datetime.now())}",
            parse_mode="Markdown"
        )

    # Удаляем файл после отправки
    os.remove(filepath)

    await query.edit_message_text(
        "✅ Файл отправлен!\nПроверь чат — должен прийти Excel.",
        reply_markup=main_menu_keyboard()
    )


# ==================== НАСТРОЙКИ ====================
async def handle_settings(query, context, user_id):
    """Меню настроек"""
    keyboard = [
        [InlineKeyboardButton("🔔 Уведомления", callback_data="notif_settings")],
        [InlineKeyboardButton("📊 Лимиты", callback_data="limits_settings")],
        [InlineKeyboardButton("🗑️ Очистить историю", callback_data="clear_history")],
        [InlineKeyboardButton("🔙 Назад", callback_data="back_menu")],
    ]

    await query.edit_message_text(
        "⚙️ *Настройки*\n\n"
        "Здесь можно настроить бота под себя.\n"
        "(В текущей версии настройки базовые, расширим позже)",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


# ==================== ОБРАБОТКА ТЕКСТА ====================
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка обычных сообщений"""
    await update.message.reply_text(
        "Я работаю только с кнопками! 👇\n"
        "Используй меню ниже или нажми /start",
        reply_markup=main_menu_keyboard()
    )


# ==================== ГЛАВНАЯ ФУНКЦИЯ ====================
def main():
    """Запуск бота"""
    # Запускаем Flask в отдельном потоке (для UptimeRobot)
    flask_thread = Thread(target=run_flask, daemon=True)
    flask_thread.start()
    logger.info("🌐 Flask endpoint запущен для UptimeRobot")

    # Создаём event loop для Python 3.14+
    import asyncio
    try:
        loop = asyncio.get_event_loop()
    except RuntimeError:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

    # Логируем запуск
    logger.info("🚀 Запуск бота...")
    logger.info(f"📁 Рабочая директория: {os.getcwd()}")
    logger.info(f"📁 База данных: {DB_PATH}")

    # Инициализируем базу
    init_db()

    # Создаём приложение
    try:
        application = Application.builder().token(TOKEN).build()
    except Exception as e:
        logger.error(f"❌ Ошибка создания приложения: {e}")
        raise

    # Обработчики команд
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_command))

    # Обработчик кнопок
    application.add_handler(CallbackQueryHandler(button_handler))

    # Обработчик текста
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    # Запускаем бота
    logger.info("✅ Бот запущен! Ожидаю сообщения...")
    try:
        application.run_polling(allowed_updates=Update.ALL_TYPES, drop_pending_updates=True)
    except Exception as e:
        logger.error(f"❌ Ошибка запуска polling: {e}")
        raise


if __name__ == "__main__":
    main()
